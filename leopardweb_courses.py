#!/usr/bin/env python3
"""
LeopardWeb Course Catalog Fetcher

Fetches all courses for a given semester from WIT's LeopardWeb system.
By default, fetches detailed meeting times and faculty information for each course.

Author: Jasper Mayone
Copyright (c) 2025 Jasper Mayone
Based on: https://github.com/WITCodingClub/calendar-backend

Usage:
    python leopardweb_courses.py <term_code>
    python leopardweb_courses.py 202510  # Spring 2025 (Excel, with details)
    python leopardweb_courses.py 202510 --format csv  # CSV output
    python leopardweb_courses.py 202510 --quick  # Fast mode (skip details)
    python leopardweb_courses.py --list-terms  # Show available terms

Output:
    Saves courses to courses_{term_code}.xlsx (default), .csv, or .json
"""

import argparse
import csv
import json
import sys
import time
from typing import Dict, List, Optional
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from colorama import Fore, Style, init

# Initialize colorama for cross-platform colored output
init(autoreset=True)


class LeopardWebClient:
    """Client for interacting with WIT's LeopardWeb system."""

    BASE_URL = "https://selfservice.wit.edu/StudentRegistrationSsb/ssb"

    def __init__(self):
        self.session = requests.Session()
        self.session_cookie: Optional[str] = None

    def get_available_terms(self) -> List[Dict]:
        """Fetch list of available academic terms."""
        url = f"{self.BASE_URL}/classSearch/getTerms"
        params = {
            "searchTerm": "",
            "offset": 1,
            "max": 50
        }

        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            terms = response.json()

            return [
                {
                    "code": term["code"],
                    "description": term["description"]
                }
                for term in terms
            ]
        except requests.RequestException as e:
            print(f"{Fore.RED}Error fetching terms: {e}", file=sys.stderr)
            return []

    def initialize_search_session(self, term: str) -> str:
        """
        Initialize a search session by POSTing term selection.
        This creates a JSESSIONID cookie that allows subsequent searches.

        Args:
            term: The term code (e.g., "202510" for Spring 2025)

        Returns:
            The JSESSIONID cookie value

        Raises:
            Exception: If session initialization fails
        """
        url = f"{self.BASE_URL}/term/search"
        params = {"mode": "search"}
        data = f"term={term}"

        headers = {
            "Content-Type": "application/x-www-form-urlencoded"
        }

        try:
            response = self.session.post(url, params=params, data=data, headers=headers)
            response.raise_for_status()

            # Extract JSESSIONID from cookies
            jsessionid = response.cookies.get("JSESSIONID")
            if not jsessionid:
                raise Exception("Failed to obtain session cookie")

            self.session_cookie = jsessionid
            return jsessionid

        except requests.RequestException as e:
            raise Exception(f"Failed to initialize search session: {e}")

    def fetch_catalog_page(self, term: str, offset: int, page_size: int = 500) -> Dict:
        """
        Fetch a single page of course catalog results.

        Args:
            term: The term code
            offset: Starting offset for pagination
            page_size: Number of results per page

        Returns:
            Dictionary containing course data and pagination info
        """
        if not self.session_cookie:
            raise Exception("Session not initialized - call initialize_search_session first")

        url = f"{self.BASE_URL}/searchResults/searchResults"

        # Generate unique session ID (mimics browser behavior)
        unique_session_id = f"sess{int(time.time())}{int(time.time() * 1000) % 10000}"

        params = {
            "txt_term": term,
            "startDatepicker": "",
            "endDatepicker": "",
            "uniqueSessionId": unique_session_id,
            "pageOffset": offset,
            "pageMaxSize": page_size,
            "sortColumn": "subjectDescription",
            "sortDirection": "asc"
        }

        headers = {
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "Accept-Language": "en-US,en;q=0.9",
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://selfservice.wit.edu/StudentRegistrationSsb/ssb/courseSearch/courseSearch",
            "Cookie": f"JSESSIONID={self.session_cookie}"
        }

        try:
            response = self.session.get(url, params=params, headers=headers)
            response.raise_for_status()
            return response.json()

        except requests.RequestException as e:
            raise Exception(f"Failed to fetch catalog page: {e}")

    def get_class_details(self, term: str, crn: str) -> Optional[Dict]:
        """
        Fetch detailed information for a specific course.

        Args:
            term: The term code
            crn: Course reference number

        Returns:
            Detailed course information including meeting times
        """
        url = f"{self.BASE_URL}/searchResults/getClassDetails"
        params = {
            "term": term,
            "courseReferenceNumber": crn
        }

        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json() if response.text else None
        except requests.RequestException:
            return None

    def get_faculty_meeting_times(self, term: str, crn: str) -> Optional[Dict]:
        """
        Fetch faculty meeting times for a specific course.

        Args:
            term: The term code
            crn: Course reference number

        Returns:
            Meeting times data
        """
        url = f"{self.BASE_URL}/searchResults/getFacultyMeetingTimes"
        params = {
            "term": term,
            "courseReferenceNumber": crn
        }

        try:
            response = self.session.get(url, params=params)
            response.raise_for_status()
            return response.json() if response.text else None
        except requests.RequestException:
            return None

    def get_course_catalog(self, term: str, detailed: bool = False, verbose: bool = True) -> List[Dict]:
        """
        Fetch all courses for a given term with pagination.

        Args:
            term: The term code (e.g., "202510")
            detailed: If True, fetch detailed info for each course (slower)
            verbose: Whether to print progress messages

        Returns:
            List of all course dictionaries
        """
        # Initialize session
        if verbose:
            print(f"{Fore.CYAN}🔄 Initializing search session for term {term}...")
        self.initialize_search_session(term)

        all_courses = []
        offset = 0
        page_size = 500
        total_count = None

        if verbose:
            print(f"{Fore.CYAN}📚 Fetching course catalog...")

        while True:
            data = self.fetch_catalog_page(term, offset, page_size)

            courses = data.get("data", [])
            if total_count is None:
                total_count = data.get("totalCount", 0)
                if verbose:
                    print(f"{Fore.YELLOW}📊 Total courses to fetch: {Style.BRIGHT}{total_count}")

            all_courses.extend(courses)

            if verbose:
                print(f"{Fore.CYAN}   Fetched {len(all_courses)}/{total_count} courses...")

            # Break if we've fetched all courses
            if len(all_courses) >= total_count or not courses:
                break

            offset += page_size

        if verbose:
            print(f"{Fore.GREEN}✓ Successfully fetched {Style.BRIGHT}{len(all_courses)}{Style.NORMAL} courses")

        # Fetch detailed information if requested
        if detailed and all_courses:
            if verbose:
                print(f"\n{Fore.CYAN}🔍 Fetching detailed information for {len(all_courses)} courses...")
                print(f"{Fore.YELLOW}⏱️  This may take a few minutes...")

            # Use tqdm progress bar with color
            course_iterator = tqdm(
                all_courses,
                desc=f"{Fore.CYAN}Fetching details",
                unit="course",
                bar_format='{l_bar}{bar:30}{r_bar}',
                colour='green'
            ) if verbose else all_courses

            for course in course_iterator:
                crn = course.get("courseReferenceNumber")
                if crn:
                    # Get detailed class info
                    details = self.get_class_details(term, crn)
                    if details:
                        course["_detailed"] = details

                    # Get faculty meeting times
                    meeting_times = self.get_faculty_meeting_times(term, crn)
                    if meeting_times and meeting_times.get("fmt"):
                        course["_faculty_meeting_times"] = meeting_times["fmt"]

            if verbose:
                print(f"{Fore.GREEN}✓ Completed detailed fetch for all courses")

        return all_courses


def flatten_course_data(course: Dict) -> Dict:
    """
    Flatten nested course data for tabular output.

    Args:
        course: Raw course data from API

    Returns:
        Flattened dictionary suitable for CSV/Excel
    """
    # Extract faculty names
    faculty_names = []
    if course.get("faculty"):
        for faculty in course["faculty"]:
            display_name = faculty.get("displayName", "")
            if display_name:
                faculty_names.append(display_name)

    # Extract meeting times - use detailed data if available
    meeting_days = []
    meeting_times = []
    meeting_locations = []

    # Try to use detailed faculty meeting times first
    if course.get("_faculty_meeting_times"):
        for fmt_data in course["_faculty_meeting_times"]:
            mt = fmt_data.get("meetingTime", {})

            # Days
            days = []
            for day in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]:
                if mt.get(day):
                    days.append(day[0].upper())
            if days:
                meeting_days.append("".join(days))

            # Times
            begin_time = mt.get("beginTime", "")
            end_time = mt.get("endTime", "")
            if begin_time and end_time:
                meeting_times.append(f"{begin_time}-{end_time}")

            # Location
            building = mt.get("building", "")
            building_desc = mt.get("buildingDescription", "")
            room = mt.get("room", "")

            location = building_desc if building_desc else building
            if room:
                location = f"{location} {room}".strip()
            if location:
                meeting_locations.append(location)

    # Fall back to regular meeting data if no detailed data
    elif course.get("meetingsFaculty"):
        for meeting in course["meetingsFaculty"]:
            mt = meeting.get("meetingTime", {})

            # Days
            days = []
            for day in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]:
                if mt.get(day):
                    days.append(day[0].upper())
            if days:
                meeting_days.append("".join(days))

            # Times
            begin_time = mt.get("beginTime", "")
            end_time = mt.get("endTime", "")
            if begin_time and end_time:
                meeting_times.append(f"{begin_time}-{end_time}")

            # Location
            building = mt.get("building", "")
            room = mt.get("room", "")
            if building or room:
                meeting_locations.append(f"{building} {room}".strip())

    return {
        "CRN": course.get("courseReferenceNumber", ""),
        "Subject": course.get("subject", ""),
        "Course Number": course.get("courseNumber", ""),
        "Section": course.get("sequenceNumber", ""),
        "Title": course.get("courseTitle", ""),
        "Credit Hours": course.get("creditHours", ""),
        "Schedule Type": course.get("scheduleTypeDescription", ""),
        "Instructional Method": course.get("instructionalMethod", ""),
        "Faculty": ", ".join(faculty_names) if faculty_names else "",
        "Meeting Days": ", ".join(meeting_days) if meeting_days else "",
        "Meeting Times": ", ".join(meeting_times) if meeting_times else "",
        "Location": ", ".join(meeting_locations) if meeting_locations else "",
        "Campus": course.get("campusDescription", ""),
        "Enrollment Current": course.get("enrollment", ""),
        "Enrollment Max": course.get("maximumEnrollment", ""),
        "Seats Available": course.get("seatsAvailable", ""),
        "Waitlist Current": course.get("waitCount", ""),
        "Waitlist Max": course.get("waitCapacity", ""),
    }


def save_as_excel(courses: List[Dict], term: str, output_file: str, verbose: bool = True):
    """Save courses to Excel workbook with formatting."""
    if verbose:
        print(f"{Fore.CYAN}📊 Creating Excel workbook...")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Courses {term}"

    # Flatten course data
    flattened_courses = [flatten_course_data(course) for course in courses]

    if not flattened_courses:
        if verbose:
            print(f"{Fore.YELLOW}⚠️  No courses to save")
        return

    # Get headers from first course
    headers = list(flattened_courses[0].keys())

    # Write headers with formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_num, course in enumerate(flattened_courses, 2):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=course.get(header, ""))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(str(header))

        for row_num in range(2, min(102, ws.max_row + 1)):  # Check first 100 rows
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set width with reasonable limits
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save workbook
    wb.save(output_file)

    if verbose:
        print(f"{Fore.GREEN}✓ Saved {Style.BRIGHT}{len(courses)}{Style.NORMAL} courses to {Style.BRIGHT}{output_file}")


def save_as_csv(courses: List[Dict], term: str, output_file: str, verbose: bool = True):
    """Save courses to CSV file."""
    if verbose:
        print(f"{Fore.CYAN}📄 Creating CSV file...")

    flattened_courses = [flatten_course_data(course) for course in courses]

    if not flattened_courses:
        if verbose:
            print(f"{Fore.YELLOW}⚠️  No courses to save")
        return

    headers = list(flattened_courses[0].keys())

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(flattened_courses)

    if verbose:
        print(f"{Fore.GREEN}✓ Saved {Style.BRIGHT}{len(courses)}{Style.NORMAL} courses to {Style.BRIGHT}{output_file}")


def save_as_json(courses: List[Dict], term: str, output_file: str, verbose: bool = True):
    """Save courses to JSON file."""
    if verbose:
        print(f"{Fore.CYAN}📝 Creating JSON file...")

    with open(output_file, 'w') as f:
        json.dump({
            "term": term,
            "total_count": len(courses),
            "courses": courses
        }, f, indent=2)

    if verbose:
        print(f"{Fore.GREEN}✓ Saved {Style.BRIGHT}{len(courses)}{Style.NORMAL} courses to {Style.BRIGHT}{output_file}")


def list_terms():
    """List all available terms."""
    client = LeopardWebClient()
    terms = client.get_available_terms()

    if not terms:
        print(f"{Fore.RED}No terms found or error occurred", file=sys.stderr)
        return

    print(f"\n{Fore.CYAN}{Style.BRIGHT}Available Terms:")
    print(f"{Fore.CYAN}" + "-" * 60)
    for term in terms:
        print(f"{Fore.YELLOW}{term['code']:10} {Fore.WHITE}{term['description']}")
    print()


def fetch_courses(term: str, output_file: Optional[str] = None,
                  format: str = "excel", quick: bool = False, verbose: bool = True):
    """
    Fetch and save courses for a given term.

    Args:
        term: The term code
        output_file: Optional custom output filename
        format: Output format ("excel", "csv", or "json")
        quick: If True, skip detailed fetch for faster execution
        verbose: Whether to print progress
    """
    client = LeopardWebClient()

    try:
        # Fetch detailed data by default, unless quick mode is enabled
        courses = client.get_course_catalog(term, detailed=not quick, verbose=verbose)

        # Determine output filename based on format
        if not output_file:
            extensions = {"excel": ".xlsx", "csv": ".csv", "json": ".json"}
            output_file = f"courses_{term}{extensions.get(format, '.xlsx')}"

        # Save in requested format
        if format == "excel":
            save_as_excel(courses, term, output_file, verbose)
        elif format == "csv":
            save_as_csv(courses, term, output_file, verbose)
        elif format == "json":
            save_as_json(courses, term, output_file, verbose)
        else:
            raise ValueError(f"Unsupported format: {format}")

    except Exception as e:
        print(f"{Fore.RED}❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Fetch course catalog from WIT LeopardWeb",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # List available terms
  python leopardweb_courses.py --list-terms

  # Fetch courses for Spring 2025 (Excel, with detailed info)
  python leopardweb_courses.py 202510

  # Fetch as CSV
  python leopardweb_courses.py 202510 --format csv

  # Fast mode (skip detailed fetch)
  python leopardweb_courses.py 202510 --quick

  # Fetch as JSON with custom output file
  python leopardweb_courses.py 202510 --format json -o spring2025.json
        """
    )

    parser.add_argument(
        "term",
        nargs="?",
        help="Term code (e.g., 202510 for Spring 2025)"
    )

    parser.add_argument(
        "--list-terms",
        action="store_true",
        help="List all available terms"
    )

    parser.add_argument(
        "-f", "--format",
        choices=["excel", "csv", "json"],
        default="excel",
        help="Output format (default: excel)"
    )

    parser.add_argument(
        "-o", "--output",
        help="Output filename (default: courses_{term}.{ext})"
    )

    parser.add_argument(
        "--quick",
        action="store_true",
        help="Skip detailed fetch for faster execution (less complete data)"
    )

    parser.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="Suppress progress messages"
    )

    args = parser.parse_args()

    # Show help if no arguments
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)

    # List terms mode
    if args.list_terms:
        list_terms()
        return

    # Fetch courses mode
    if not args.term:
        print("Error: term code is required (or use --list-terms)", file=sys.stderr)
        parser.print_help()
        sys.exit(1)

    fetch_courses(args.term, args.output, args.format, quick=args.quick, verbose=not args.quiet)


if __name__ == "__main__":
    main()
